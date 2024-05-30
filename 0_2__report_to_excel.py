import sqlite3
import openpyxl

# Connect to the database
conn = sqlite3.connect("excel_data5.db")
c = conn.cursor()

# Create a new Excel workbook
workbook = openpyxl.Workbook()

# 1. Export "RF_NEW_DATA" to the "RF_NEW_DATA" sheet
worksheet = workbook.active
worksheet.title = "RF_NEW_DATA"

# Get table headers
c.execute("SELECT * FROM RF_NEW_DATA")
headers = [desc[0] for desc in c.description]
worksheet.append(headers)

# Get table data
rows = c.fetchall()
for row in rows:
    worksheet.append(row)

# 2. Export "excel_rf_old_data" to the "RF_OLD_DATA" sheet
worksheet = workbook.create_sheet("RF_OLD_DATA")

# Get table headers
c.execute("SELECT * FROM excel_rf_old_data")
headers = [desc[0] for desc in c.description]
worksheet.append(headers)

# Get table data
rows = c.fetchall()
for row in rows:
    worksheet.append(row)

# 3. Export combined "Site_id_information" and "TSS_date" to the "Site_Information" sheet
worksheet = workbook.create_sheet("Site_Information")

# Get table headers from both tables
c.execute("SELECT * FROM Site_id_information")
headers1 = [desc[0] for desc in c.description]
c.execute("SELECT * FROM TSS_date")
headers2 = [desc[0] for desc in c.description]
headers = headers1 + headers2
worksheet.append(headers)

# Combine and write data from both tables
rows1 = c.execute("SELECT * FROM Site_id_information").fetchall()
rows2 = c.execute("SELECT * FROM TSS_date").fetchall()
rows = [row1 + row2 for row1, row2 in zip(rows1, rows2)]
for row in rows:
    worksheet.append(row)

# Save the Excel file
workbook.save("TSSR report data.xlsx")

# Close the database connection
conn.close()

print("Report exported successfully!")
