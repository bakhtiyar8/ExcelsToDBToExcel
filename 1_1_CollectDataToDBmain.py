import os
import openpyxl
import sqlite3

# Путь к папке с файлами Excel
folder_path = "./site TSSR/exc/"
print(folder_path)

# Создаем базу данных SQLite и подключаемся к ней
conn = sqlite3.connect('excel_data5.db')
c = conn.cursor()

# Создаем таблицу в базе данных для хранения данных из Excel
# Adding print statements to verify table creation
print("Creating tables...")

c.execute('''CREATE TABLE IF NOT EXISTS TSS_date (
    "TSS date" TEXT,
    "RF Planner" TEXT,
    "TR Planner" TEXT,
    "Тип работ" TEXT,
    "Old config BTS" TEXT,
    "Final config BTS" TEXT,
    "Высота над уровнем моря" TEXT
);''')
print("Created table TSS_date")

c.execute('''CREATE TABLE IF NOT EXISTS Site_id_information (
    "Site id" TEXT,
    "Site adress / Адрес" TEXT,
    "Coordinates dec / Координаты: Long." TEXT,
    "Coordinates dec / Координаты: Lat." TEXT,
    "Tower Type/ Тип мачты" TEXT,
    "Length of Tower/Высота мачты" TEXT,
    "Height Rooftop (m)/ высота крыши (м)" TEXT,
    "Summ RF Ant. / Cуммарное количество RF антенн" TEXT,
    "Equipment type of rack| Тип БС indoor/outdoor" TEXT,
    "Equipment type" TEXT,
    "Power Equipment type/ Термошкаф" TEXT
);''')
print("Created table Site_id_information")

c.execute('''CREATE TABLE IF NOT EXISTS excel_rf_old_data (
    "RF OLD DATA: Site id" TEXT,
    "RF OLD DATA: Cell name" TEXT,
    "RF OLD DATA: Cell number" TEXT,
    "RF OLD DATA: Band / Диапазон" TEXT,
    "RF OLD DATA: Antenna height / Высота подвеса антенн" TEXT,
    "RF OLD DATA: Antenna / Антенна" TEXT,
    "RF OLD DATA: Azimuts / Азимуты" TEXT,
    "RF OLD DATA: Mech.tilt " TEXT,
    "RF OLD DATA: Electr.tilt" TEXT,
    "RF OLD DATA: RRU" TEXT,
    "RF OLD DATA: Сombainer/TMA" TEXT
);''')
print("Created table excel_rf_old_data")

c.execute('''CREATE TABLE IF NOT EXISTS RF_NEW_DATA (
    "RF NEW DATA: Site id" TEXT,
    "RF NEW DATA: Cell name" TEXT,
    "RF NEW DATA: Cell number" TEXT,
    "RF NEW DATA: Band / Диапазон" TEXT,
    "RF NEW DATA: Antenna height / Высота подвеса антенн" TEXT,
    "RF NEW DATA: Antenna / Антенна" TEXT,
    "RF NEW DATA: Azimuts / Азимуты" TEXT,
    "RF NEW DATA: Mech.tilt " TEXT,
    "RF NEW DATA: Electr.tilt" TEXT,
    "RF NEW DATA: RRU" TEXT,
    "RF NEW DATA: Сombainer/TMA" TEXT
);''')
print("Created table RF_NEW_DATA")

print("Tables created successfully.")

# Проходим по всем файлам в папке
for file_name in os.listdir(folder_path):
    if file_name.endswith(".xlsx"):
        print(f"Processing file: {file_name}")
        # Открываем файл Excel
        wb = openpyxl.load_workbook(os.path.join(folder_path, file_name))
        target_partial_name = 'SITE PARAM'
        for sheet_name in wb.sheetnames:
            print(f"Checking sheet: {sheet_name}")
            if target_partial_name.lower() in sheet_name.lower():
                selected_sheet = wb[sheet_name]
                break
        else:
            print(f"No sheet with a partial name '{target_partial_name}' exists in the workbook.")
            continue

        # Получаем данные из таблицы начиная с A3 по J3
        for row in range(3, 4):
            data = [selected_sheet.cell(row=row, column=col).value for col in range(1, 8)]
            print(f"Row 3 data: {data}")
            c.execute("INSERT INTO TSS_date VALUES (?, ?, ?, ?, ?, ?, ?)", data)

        # Получаем данные из таблицы начиная с A8 по L8
        for row in range(8, 9):
            data = [selected_sheet.cell(row=row, column=col).value for col in range(1, 12)]
            print(f"Row 8 data: {data}")
            c.execute("INSERT INTO Site_id_information VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)", data)

        # Получаем данные из таблицы начиная с A13 по F13
        for row in range(13, 14):
            data = [selected_sheet.cell(row=row, column=col).value for col in range(1, 12)]
            print(f"Row 13 data: {data}")
            c.execute("INSERT INTO excel_rf_old_data VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)", data)

        # Получаем данные построчно до пустой строки
        row = 14
        while selected_sheet.cell(row=row, column=1).value is not None:
            data = [selected_sheet.cell(row=row, column=col).value for col in range(1, 12)]
            print(f"Row {row} data: {data}")
            c.execute("INSERT INTO excel_rf_old_data VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)", data)
            row += 1

        # RF NEW DATA
        rf_new_data = []
        for row in range(1, selected_sheet.max_row + 1):
            if selected_sheet.cell(row=row, column=1).value == "RF NEW DATA":
                rf_new_data_start_row = row + 2
                break

        if rf_new_data_start_row:
            row = rf_new_data_start_row
            while selected_sheet.cell(row=row, column=1).value is not None:
                data = tuple(selected_sheet.cell(row=row, column=i).value for i in range(1, 12))
                print(f"RF NEW DATA row {row}: {data}")
                rf_new_data.append(data)
                row += 1
            c.executemany("INSERT INTO RF_NEW_DATA VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)", rf_new_data)

# Сохраняем изменения в базе данных и закрываем соединение
conn.commit()
conn.close()
